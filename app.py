from flask import Flask, render_template, request, jsonify, send_file
import requests
import openpyxl
import io
import time
import json

app = Flask(__name__)

NPI_API_URL = "https://npiregistry.cms.hhs.gov/api/"


def format_address(addr: dict) -> str:
    if not addr:
        return ""
    parts = [
        addr.get("address_1", ""),
        addr.get("address_2", ""),
        addr.get("city", ""),
        addr.get("state", ""),
        addr.get("postal_code", ""),
    ]
    country = addr.get("country_name", "")
    if country and country != "United States":
        parts.append(country)
    phone = addr.get("telephone_number", "")
    address_str = ", ".join(p for p in parts if p)
    if phone:
        address_str += f" | Phone: {phone}"
    return address_str


def lookup_npi(npi: str) -> dict:
    empty = {
        "npi": "", "enumeration_date": "", "npi_type": "", "status": "",
        "provider_name": "", "authorized_official": "", "mailing_address": "",
        "primary_practice_address": "", "secondary_practice_addresses": "",
        "health_information_exchange": "", "other_identifiers": "", "taxonomy": "",
        "error": ""
    }
    try:
        params = {"number": npi.strip(), "version": "2.1"}
        response = requests.get(NPI_API_URL, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()

        results = data.get("results", [])
        if not results:
            empty["error"] = "Not found"
            return empty

        r = results[0]
        basic = r.get("basic", {})

        empty["npi"] = r.get("number", "")
        empty["enumeration_date"] = basic.get("enumeration_date", "")

        npi_type_code = r.get("enumeration_type", "")
        empty["npi_type"] = "NPI-1 Individual" if npi_type_code == "NPI-1" else "NPI-2 Organization" if npi_type_code == "NPI-2" else npi_type_code
        empty["status"] = basic.get("status", "")

        if npi_type_code == "NPI-2":
            name = basic.get("organization_name", "")
        else:
            name_parts = [basic.get("name_prefix",""), basic.get("first_name",""),
                          basic.get("middle_name",""), basic.get("last_name",""),
                          basic.get("name_suffix",""), basic.get("credential","")]
            name = " ".join(p for p in name_parts if p)
        empty["provider_name"] = name

        ao_first = basic.get("authorized_official_first_name", "")
        ao_last  = basic.get("authorized_official_last_name", "")
        if ao_first or ao_last:
            ao_parts = [basic.get("authorized_official_name_prefix",""), ao_first,
                        basic.get("authorized_official_middle_name",""), ao_last]
            ao_name  = " ".join(p for p in ao_parts if p)
            ao_title = basic.get("authorized_official_title_or_position", "")
            ao_phone = basic.get("authorized_official_telephone_number", "")
            ao_str   = f"Name: {ao_name}"
            if ao_title: ao_str += f" | Title: {ao_title}"
            if ao_phone: ao_str += f" | Phone: {ao_phone}"
            empty["authorized_official"] = ao_str

        addresses = r.get("addresses", [])
        mailing  = next((a for a in addresses if a.get("address_purpose") == "MAILING"),  None)
        location = next((a for a in addresses if a.get("address_purpose") == "LOCATION"), None)
        empty["mailing_address"]          = format_address(mailing)
        empty["primary_practice_address"] = format_address(location)

        practice_locations = r.get("practiceLocations", [])
        if practice_locations:
            empty["secondary_practice_addresses"] = " | ".join(format_address(pl) for pl in practice_locations)

        endpoints = r.get("endpoints", [])
        if endpoints:
            hie_parts = []
            for ep in endpoints:
                ep_str = " | ".join(p for p in [ep.get("endpointType",""), ep.get("endpoint",""),
                    ep.get("endpointDescription",""), ep.get("use",""), ep.get("contentType",""),
                    ep.get("affiliation","")] if p)
                if ep_str: hie_parts.append(ep_str)
            empty["health_information_exchange"] = "; ".join(hie_parts)

        identifiers = r.get("identifiers", [])
        if identifiers:
            id_parts = []
            for ident in identifiers:
                id_str = " | ".join(p for p in [ident.get("desc",""), ident.get("identifier",""),
                    ident.get("state",""), ident.get("issuer","")] if p)
                if id_str: id_parts.append(id_str)
            empty["other_identifiers"] = "; ".join(id_parts)

        taxonomies = r.get("taxonomies", [])
        if taxonomies:
            tax_parts = []
            for tax in taxonomies:
                tax_str = f"{tax.get('desc','')} [{tax.get('code','')}]"
                if tax.get("primary"): tax_str += " (Primary)"
                if tax.get("state"):   tax_str += f" | State: {tax.get('state')}"
                if tax.get("license"): tax_str += f" | License: {tax.get('license')}"
                tax_parts.append(tax_str)
            empty["taxonomy"] = "; ".join(tax_parts)

        return empty

    except Exception as e:
        empty["error"] = str(e)
        return empty


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/lookup-single", methods=["POST"])
def lookup_single():
    data = request.get_json()
    npi = data.get("npi", "").strip()
    if not npi:
        return jsonify({"error": "No NPI provided"}), 400
    result = lookup_npi(npi)
    return jsonify(result)


@app.route("/lookup-batch", methods=["POST"])
def lookup_batch():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        return jsonify({"error": "Please upload an .xlsx, .xls, or .csv file"}), 400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        npi_rows = []
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            cell = row[0]
            val = str(cell.value).strip() if cell.value is not None else ""
            if val and val != "None":
                npi_rows.append((cell.row, val))

        if not npi_rows:
            return jsonify({"error": "No NPI numbers found in Column A starting from row 2"}), 400

        # Write headers
        headers = [
            "NPI", "NPI (Registry Confirmed)", "Enumeration Date", "NPI Type",
            "Status", "Provider Name", "Authorized Official Information",
            "Mailing Address", "Primary Practice Address",
            "Secondary Practice Address(es)", "Health Information Exchange",
            "Other Identifiers", "Taxonomy"
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        results_log = []

        for row_num, npi in npi_rows:
            info = lookup_npi(npi)
            ws.cell(row=row_num, column=1,  value=npi)
            ws.cell(row=row_num, column=2,  value=info["npi"])
            ws.cell(row=row_num, column=3,  value=info["enumeration_date"])
            ws.cell(row=row_num, column=4,  value=info["npi_type"])
            ws.cell(row=row_num, column=5,  value=info["status"])
            ws.cell(row=row_num, column=6,  value=info["provider_name"])
            ws.cell(row=row_num, column=7,  value=info["authorized_official"])
            ws.cell(row=row_num, column=8,  value=info["mailing_address"])
            ws.cell(row=row_num, column=9,  value=info["primary_practice_address"])
            ws.cell(row=row_num, column=10, value=info["secondary_practice_addresses"])
            ws.cell(row=row_num, column=11, value=info["health_information_exchange"])
            ws.cell(row=row_num, column=12, value=info["other_identifiers"])
            ws.cell(row=row_num, column=13, value=info["taxonomy"])

            results_log.append({
                "npi": npi,
                "provider_name": info["provider_name"],
                "primary_practice_address": info["primary_practice_address"],
                "status": info["status"],
                "error": info["error"]
            })
            time.sleep(0.3)

        # Save to in-memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="NPI_List_Full_Data.xlsx"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
